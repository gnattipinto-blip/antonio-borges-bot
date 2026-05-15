import os
import openpyxl
from openpyxl.styles import PatternFill

PLANILHA = os.path.join(os.path.dirname(__file__), "Unidades_AntonioBorges.xlsx")

CORES = {
    "DISPONÍVEL": PatternFill("solid", fgColor="C6EFCE"),
    "RESERVADO":  PatternFill("solid", fgColor="FFEB9C"),
    "VENDIDO":    PatternFill("solid", fgColor="FFC7CE"),
}

def _carregar():
    return openpyxl.load_workbook(PLANILHA)

def _salvar(wb):
    wb.save(PLANILHA)

def _get_rows(ws):
    rows = []
    for row in ws.iter_rows(min_row=3, values_only=False):
        val = row[0].value
        if val and str(val).startswith("Apto"):
            rows.append({
                "row_idx": row[0].row,
                "unidade":   str(row[0].value or "").strip(),
                "bloco":     str(row[1].value or "").strip(),
                "tipo":      str(row[2].value or "").strip(),
                "status":    str(row[3].value or "").strip().upper(),
                "comprador": str(row[4].value or "").strip(),
                "obs":       str(row[5].value or "").strip(),
            })
    return rows

def listar_disponiveis():
    wb = _carregar()
    ws = wb.active
    return [r for r in _get_rows(ws) if r["status"] == "DISPONÍVEL"]

def listar_todas():
    wb = _carregar()
    ws = wb.active
    return _get_rows(ws)

def resumo():
    todos = listar_todas()
    total = len(todos)
    disp = sum(1 for u in todos if u["status"] == "DISPONÍVEL")
    res  = sum(1 for u in todos if u["status"] == "RESERVADO")
    vend = sum(1 for u in todos if u["status"] == "VENDIDO")
    pct  = ((res + vend) / total * 100) if total else 0
    return {"total": total, "disponiveis": disp, "reservadas": res,
            "vendidas": vend, "pct_ocupado": pct}

def _atualizar_status(unidade, novo_status, comprador=""):
    unidade = unidade.strip()
    wb = _carregar()
    ws = wb.active
    rows = _get_rows(ws)
    encontrada = None
    for r in rows:
        if r["unidade"].upper() == unidade.upper():
            encontrada = r
            break
    if not encontrada:
        return False, f"❌ Unidade *{unidade}* não encontrada."
    row_idx = encontrada["row_idx"]
    ws.cell(row=row_idx, column=4).value = novo_status
    ws.cell(row=row_idx, column=4).fill = CORES[novo_status]
    ws.cell(row=row_idx, column=5).value = comprador
    _salvar(wb)
    return True, None

def marcar_reservado(unidade, comprador):
    ok, err = _atualizar_status(unidade, "RESERVADO", comprador)
    if not ok:
        return False, err
    return True, f"🔒 *{unidade}* marcada como RESERVADO para *{comprador}*."

def marcar_vendido(unidade, comprador):
    ok, err = _atualizar_status(unidade, "VENDIDO", comprador)
    if not ok:
        return False, err
    return True, f"✅ *{unidade}* marcada como VENDIDO para *{comprador}*."

def marcar_disponivel(unidade):
    ok, err = _atualizar_status(unidade, "DISPONÍVEL", "")
    if not ok:
        return False, err
    return True, f"🔓 *{unidade}* está DISPONÍVEL novamente."
